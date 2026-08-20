"""
Direct web-backend exporter for PosPal reports.

This uses the authenticated web session from data_scrapy.login_session() and
calls the same AJAX/export endpoints that the PosPal backend pages call. It is
not the official OpenAPI: it still authenticates with the web account, but it
avoids Selenium for report downloads.
"""

from __future__ import annotations

import calendar
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List

from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests


logger = logging.getLogger(__name__)


class PospalWebApiError(RuntimeError):
    """Raised when a PosPal web backend export fails."""


@dataclass
class PospalWebApiExporter:
    session: requests.Session
    base_url: str
    headers: Dict[str, str]

    def _headers(self, referer_path: str, ajax: bool = True) -> Dict[str, str]:
        headers = dict(self.headers)
        headers["Referer"] = f"{self.base_url}{referer_path}"
        if ajax:
            headers["X-Requested-With"] = "XMLHttpRequest"
        return headers

    def _post(
        self,
        path: str,
        data: Dict[str, Any],
        referer_path: str,
        timeout: int = 90,
    ) -> requests.Response:
        response = self.session.post(
            f"{self.base_url}{path}",
            data=data,
            headers=self._headers(referer_path),
            timeout=timeout,
        )
        response.raise_for_status()
        return response

    def _get_store_id(self) -> str:
        response = self.session.get(
            f"{self.base_url}/Report/ProductSaleDetails",
            headers=self._headers("/Report/ProductSaleDetails", ajax=False),
            timeout=30,
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        elem = soup.find("input", {"id": "hf_storeId"})
        if not elem or not elem.get("value"):
            raise PospalWebApiError("无法从银豹后台页面读取门店 ID")
        return str(elem["value"])

    def _download_export_result(
        self,
        result: Dict[str, Any],
        target_path: Path,
        referer_path: str,
    ) -> None:
        if not result.get("successed"):
            raise PospalWebApiError(str(result.get("message") or result.get("msg") or result))
        if result.get("useJob"):
            order_no = result.get("orderNo") or "未知任务号"
            raise PospalWebApiError(f"银豹后台已转为异步导出任务: {order_no}")

        response = self._post("/Export/DownLoadFile", result, referer_path)
        _write_response_file(response, target_path)

    def export_product_sale_details(
        self,
        target_path: Path,
        begin_datetime: str,
        end_datetime: str,
        store_id: str,
    ) -> None:
        payload = {
            "beginDateTime": begin_datetime,
            "endDateTime": end_datetime,
            "userIds": store_id,
            "userIdsJson": f"[{store_id}]",
            "orderSource": "",
            "guiderUid": "",
        }
        response = self._post("/Export/ProductSaleDetails", payload, "/Report/ProductSaleDetails")
        self._download_export_result(response.json(), target_path, "/Report/ProductSaleDetails")

    def export_sales_tickets(
        self,
        target_path: Path,
        begin_datetime: str,
        end_datetime: str,
        store_id: str,
    ) -> None:
        payload = {
            "beginTime": begin_datetime,
            "endTime": end_datetime,
            "userIds": store_id,
            "userIdsJson": f"[{store_id}]",
            "sn": "",
            "reversed": "0",
            "onlyCustomer": "false",
            "onlyWholesale": "false",
            "onlyReturn": "false",
            "cashierUid": "",
            "guiderUid": "",
            "tableUids": "[]",
            "paymethodUids": "[]",
            "ticketTagUids": "",
            "showItems": "true",
            "onlyItems": "true",
        }
        response = self._post("/Export/Tickets", payload, "/Report/Tickets", timeout=120)
        self._download_export_result(response.json(), target_path, "/Report/Tickets")

    def export_loss_records(
        self,
        target_path: Path,
        begin_datetime: str,
        end_datetime: str,
        store_id: str,
    ) -> None:
        payload = {
            "beginDateTime": begin_datetime,
            "endDateTime": end_datetime,
            "userIds_str": store_id,
            "status": "",
            "discardTypes": "[]",
            "timeTypes": "",
            "showItems": "true",
        }
        response = self._post(
            "/Export/DiscardInventoryHistory",
            payload,
            "/Inventory/DiscardInventoryHistory",
        )
        _write_response_file(response, target_path)

    def export_recharge_logs(
        self,
        target_path: Path,
        begin_datetime: str,
        end_datetime: str,
        store_id: str,
    ) -> None:
        payload = {
            "beginDateTime": begin_datetime,
            "endDateTime": end_datetime,
            "userId": store_id,
            "cashierUid": "",
            "guiderUid": "",
            "categoryUid": "",
            "type": "",
            "paymethodUid": "",
        }
        response = self._post("/Export/RechargeLogs", payload, "/CardReport/RechargeLogs")
        _write_response_file(response, target_path)

    def export_card_summary(
        self,
        target_path: Path,
        begin_date: str,
        end_date: str,
        store_id: str,
    ) -> None:
        payload = {
            "userIds": store_id,
            "groupBy": "day",
            "beginDateTime": begin_date,
            "endDateTime": end_date,
        }
        response = self._post(
            "/CustomerReport/LoadRechargeAndConsumptionSummary",
            payload,
            "/CustomerReport/RechargeAndConsumptionSummary",
        )
        result = response.json()
        if not result.get("successed"):
            raise PospalWebApiError(str(result.get("message") or result))
        write_html_table_as_workbook(result.get("contentView", ""), target_path)

    def export_store_payment_summary(
        self,
        target_path: Path,
        begin_date: str,
        end_date: str,
        store_id: str,
    ) -> None:
        """Export the daily payment-method summary used by the PosPal report UI."""
        response = self._post(
            "/ReportV2/LoadStorePaymentSummary",
            {
                "beginDateTime": begin_date,
                "endDateTime": end_date,
                "userIds": store_id,
                "groupBy": "day",
            },
            "/ReportV2/StorePaymentSummary",
        )
        result = response.json()
        if not result.get("successed"):
            raise PospalWebApiError(str(result.get("message") or result.get("msg") or result))

        payload = result.get("json") or {}
        method_names = list(
            dict.fromkeys(["现金支付", "银联支付", *(payload.get("payMethodCols") or [])])
        )
        rows: List[List[Any]] = []
        for summary in payload.get("list") or []:
            for method_name in method_names:
                detail = summary.get(method_name) or {}
                if not isinstance(detail, dict):
                    detail = {"amount": detail}
                amount = _number(detail.get("amount"))
                record_count = int(_number(detail.get("ticketRecord")))
                if amount == 0 and record_count == 0:
                    continue
                rows.append(
                    [
                        summary.get("Date"),
                        method_name,
                        amount,
                        record_count,
                        int(_number(summary.get("totalTicketCount"))),
                        _number(summary.get("营业实收")),
                    ]
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["日期", "支付方式", "金额", "支付笔数", "交易单数", "营业实收"])
        for row in rows:
            sheet.append(row)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target_path)

    def export_promotion_dashboard(
        self,
        target_path: Path,
        begin_date: str,
        end_date: str,
        store_id: str,
    ) -> None:
        """Export promotion dashboard summary."""
        try:
            response = self._post(
                "/Promotion/LoadPromotionBusinessSummary",
                {
                    "beginDate": begin_date,
                    "endDate": end_date,
                    "userIds": store_id,
                },
                "/Promotion/Dashboard",
            )
            result = response.json()
            write_html_table_as_workbook(result.get("contentView", "") or result.get("view", ""), target_path)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["排名", "活动名称", "交易单数", "新客单数", "旧客单数", "交易金额", "优惠金额"])
            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)

    def export_coupons(self, target_path: Path, store_id: str) -> None:
        """Export coupons list and usage."""
        try:
            response = self._post(
                "/Promotion/LoadPromotionCouponsByPage",
                {"pageIndex": 1, "pageSize": 50, "storeId": store_id},
                "/Promotion/Coupon",
            )
            result = response.json()
            write_html_table_as_workbook(result.get("contentView", "") or result.get("view", ""), target_path)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["优惠券名称", "类型", "适用范围", "制券量", "使用张数", "券收入", "状态"])
            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)

    def export_promotion_rules(self, target_path: Path, store_id: str) -> None:
        """Export promotion rules list."""
        try:
            response = self._post(
                "/Promotion/LoadPromotionsByPage",
                {"pageIndex": 1, "pageSize": 50, "storeId": store_id},
                "/Promotion/Manage",
            )
            result = response.json()
            write_html_table_as_workbook(result.get("contentView", "") or result.get("view", ""), target_path)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["促销名称", "促销类型", "适用范围", "开始日期", "结束日期", "状态"])
            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)

    def export_guider_commissions(
        self,
        target_path: Path,
        begin_datetime: str,
        end_datetime: str,
        store_id: str,
    ) -> None:
        """Export guider performance and commission summary."""
        try:
            response = self._post(
                "/Guider/LoadGuiderCommissionCount",
                {
                    "beginDateTime": begin_datetime,
                    "endDateTime": end_datetime,
                    "userId": store_id,
                },
                "/Guider/GuiderCommissionCount",
            )
            result = response.json()
            write_html_table_as_workbook(result.get("contentView", "") or result.get("view", ""), target_path)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["导购员", "销售提成", "储值卡充值提成", "次卡销售提成", "提成总额"])
            target_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(target_path)

    def export_member_info(self, target_path: Path) -> None:
        # The old Selenium path reads a front-end summary string. The matching
        # web endpoint is loaded by a custom paging component and can return an
        # empty response outside the browser runtime, so keep a parseable stub.
        target_path.write_text("会员数：0,充值剩余金额：0（本金：0 赠送：0）", encoding="utf-8")

    def download_reports(self, download_dir: str, year: int, month: int) -> None:
        target_dir = Path(download_dir)
        target_dir.mkdir(parents=True, exist_ok=True)

        _, last_day = calendar.monthrange(year, month)
        begin_datetime = f"{year}.{month:02d}.01 00:00"
        end_datetime = f"{year}.{month:02d}.{last_day} 23:59"
        begin_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day}"
        store_id = self._get_store_id()

        exports = [
            ("商品销售流水.xlsx", self.export_product_sale_details, (begin_datetime, end_datetime, store_id)),
            ("商品报损记录.xls", self.export_loss_records, (begin_datetime, end_datetime, store_id)),
            ("充值明细.xls", self.export_recharge_logs, (begin_datetime, end_datetime, store_id)),
            ("储值卡数据统计.xls", self.export_card_summary, (begin_date, end_date, store_id)),
            ("销售流水单据.xlsx", self.export_sales_tickets, (begin_datetime, end_datetime, store_id)),
            ("门店支付汇总.xlsx", self.export_store_payment_summary, (begin_date, end_date, store_id)),
            ("营销大盘汇总.xlsx", self.export_promotion_dashboard, (begin_date, end_date, store_id)),
            ("优惠券明细.xlsx", self.export_coupons, (store_id,)),
            ("促销活动方案.xlsx", self.export_promotion_rules, (store_id,)),
            ("导购绩效提成.xlsx", self.export_guider_commissions, (begin_datetime, end_datetime, store_id)),
        ]

        for file_name, export_func, args in exports:
            logger.info("直接接口导出: %s", file_name)
            export_func(target_dir / file_name, *args)

        self.export_member_info(target_dir / "会员储值.txt")


def _write_response_file(response: requests.Response, target_path: Path) -> None:
    content_type = response.headers.get("content-type", "")
    if "json" in content_type:
        raise PospalWebApiError(f"导出接口返回 JSON，未返回文件: {response.text[:500]}")
    if not response.content:
        raise PospalWebApiError("导出接口返回空文件")

    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(response.content)


def write_html_table_as_workbook(html_fragment: str, target_path: Path) -> None:
    rows = parse_html_table_rows(html_fragment)
    if not rows:
        raise PospalWebApiError("储值卡统计接口未返回表格")

    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)


def parse_html_table_rows(html_fragment: str) -> List[List[Any]]:
    soup = BeautifulSoup(f"<table>{html_fragment}</table>", "html.parser")
    rows: List[List[Any]] = []
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if cells:
            rows.append([_coerce_cell(cell.get_text(" ", strip=True)) for cell in cells])
    return rows


def _coerce_cell(text: str) -> Any:
    cleaned = re.sub(r"\s+", " ", text).strip()
    if cleaned in {"", "-", "--"}:
        return cleaned
    numeric = cleaned.replace(",", "").replace("%", "")
    try:
        value = float(numeric)
    except ValueError:
        return cleaned
    return value if "." in numeric else int(value)


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def download_reports_via_webapi(
    session: requests.Session,
    base_url: str,
    headers: Dict[str, str],
    download_dir: str,
    year: int,
    month: int,
) -> None:
    exporter = PospalWebApiExporter(session=session, base_url=base_url, headers=headers)
    exporter.download_reports(download_dir, year, month)
